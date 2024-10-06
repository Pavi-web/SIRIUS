from tkinter import *
from tkinter import messagebox
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

root = Tk()
root.title("Feedback form")
root.geometry('700x400')  # Keep geometry simple without unnecessary offsets
root.resizable(False, False)
root.configure(bg="#326273")

# Check if the file exists
file_path = pathlib.Path('Backened_data.xlsx')
if file_path.exists():
    file = openpyxl.load_workbook(file_path)
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = "Full Name"
    sheet['B1'] = "Phone Number"
    sheet['C1'] = "Age"
    sheet['D1'] = "Gender"
    sheet['E1'] = "Address"
    file.save(file_path)  # Save the new file

def submit():
    name = nameValue.get()
    contact = contactValue.get()
    age = AgeValue.get()
    gender = gender_combobox.get()
    address = addressEntry.get(1.0, END)

    # Open the existing file and write data
    file = openpyxl.load_workbook(file_path)
    sheet = file.active
    sheet.cell(column=1, row=sheet.max_row + 1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contact)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save(file_path)

    messagebox.showinfo('Info', 'Detail added!')
    clear()  # Clear fields after submission

def clear():
    nameValue.set('')
    contactValue.set('')
    AgeValue.set('')
    addressEntry.delete(1.0, END)

# Icon
icon_image = PhotoImage(file="logo.png")
root.iconphoto(False, icon_image)

# Heading Label
Label(root, text="Please fill out this entry form:", font="Arial 13", bg="#326273", fg="#fff").place(x=20, y=20)

# Labels
Label(root, text='Name', font=23, bg="#326273", fg="#fff").place(x=50, y=80)
Label(root, text='Contact No', font=23, bg="#326273", fg="#fff").place(x=50, y=120)
Label(root, text='Age', font=23, bg="#326273", fg="#fff").place(x=50, y=160)
Label(root, text='Gender', font=23, bg="#326273", fg="#fff").place(x=50, y=200)  # Moved down
Label(root, text='Feedback', font=23, bg="#326273", fg="#fff").place(x=50, y=240)

# Entry fields
nameValue = StringVar()
contactValue = StringVar()
AgeValue = StringVar()

nameEntry = Entry(root, textvariable=nameValue, width=45, bd=2, font=20)
contactEntry = Entry(root, textvariable=contactValue, width=45, bd=2, font=20)
ageEntry = Entry(root, textvariable=AgeValue, width=45, bd=2, font=20)

# Gender Combobox (moved down)
gender_combobox = Combobox(root, values=['Male', 'Female'], font='Arial 14', state='readonly', width=14)
gender_combobox.place(x=200, y=200)
gender_combobox.set('Male')

addressEntry = Text(root, width=50, height=4, bd=2)
addressEntry.place(x=200, y=240)

nameEntry.place(x=200, y=80)
contactEntry.place(x=200, y=120)
ageEntry.place(x=200, y=160)

# Buttons with corrected placement
Button(root, text="Submit", bg="#326273", fg="white", width=15, height=2, command=submit).place(x=50, y=320)
Button(root, text="Clear", bg="#326273", fg="white", width=15, height=2, command=clear).place(x=250, y=320)
Button(root, text="Exit", bg="#326273", fg="white", width=15, height=2, command=root.destroy).place(x=450, y=320)

root.mainloop()
